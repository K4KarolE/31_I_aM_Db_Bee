import platform
import os
import sys
from datetime import date

from functions import settings              # from settings_db.json getting the info for the TARGET SHEET and MOVIE DB locations
from functions import messages

def write_sheet(titleRead, yearRead, directors, stars, genres, lengthHour, lengthMinute):
    cellnumber = 3

    settings_data = settings.open_settings()                                # opening the settings_db.json DB
    full_path_to_MoviesNewRecord = settings_data['path_movie_new_record']

    from openpyxl import load_workbook
    wb = load_workbook(full_path_to_MoviesNewRecord)
    ws = wb.active

## ADDING THE VALUES TO EXCEL
# MOVIE TITLE
    cell = 'C' + str(cellnumber)
    ws[cell].value = titleRead
    # YEAR OF RELEASE 
    cellRYear = 'E' + str(cellnumber)
    ws[cellRYear].value = yearRead

# DIRECTOR(S)
    for counter in range(0,3):
            cell_director = 'F' + str(int(cellnumber) + counter)
            try:
                    ws[cell_director].value = directors[counter]    # adding directors to the sheet (overwriting the privious ones)
            except:
                    ws[cell_director].value = None                  # removing previous values, example: the new title has 1 director, the previous had 3
                                                                    # the first will be overwritten, the 2nd, 3rd will be removed from the sheet

# STAR(S)
    for counter in range(0,3):
            cell_star = 'G' + str(int(cellnumber) + counter)
            try:
                    ws[cell_star].value = stars[counter]
            except:
                    ws[cell_star].value = None

# GENRE(S)
    genre_columns = ['H', 'I', 'J']
    for counter in range(0,3):
            cell_genre = genre_columns[counter] + str(cellnumber)   # writing the genre values horizontally (not vertically like: directors, actors)
            try:
                    ws[cell_genre].value = genres[counter]
            except:
                    ws[cell_genre].value = None

# MOVIE LENGTH
    cellLengthHour = 'Q' + str(cellnumber)
    ws[cellLengthHour].value = None                 # removing the previous value from the cell
    if lengthHour != None:
            ws[cellLengthHour].value = str(lengthHour)

    cellLengthMin = 'R' + str(cellnumber)
    ws[cellLengthMin].value = None    
    if lengthMinute != None:
            ws[cellLengthMin].value = str(lengthMinute)

# TODAY`S DATE
    today = date.today()

    day = 'K' + str(cellnumber)
    ws[day].value = str(today)[8:]

    day = 'L' + str(cellnumber)
    ws[day].value = str(today)[5:7]

    day = 'M' + str(cellnumber)
    ws[day].value = str(today)[0:4]

# HOW MANY TIMES SEEN FORMULA
    formula = '=COUNTA(M' + str(cellnumber) + ':M' + str(int(cellnumber) + 2) + ')'  # like: =COUNTA(M6965:M6967)
    day = 'N' + str(cellnumber)
    ws[day].value = formula

# 1st TIME WATCHING
    cellRFirst = 'O' + str(cellnumber)
    ws[cellRFirst].value = '1st'        

# SAVE THE SHEET
    openSheet = True
    counter = 0
    while openSheet == True:
            try:             
                wb.save(full_path_to_MoviesNewRecord)
                openSheet = False
                print('\n')
            except:
                counter += 1
                if counter < 4:
                    messages.error_pop_up('excel_is_open')
                else:
                    sys.exit()

def launch_sheets():
    if platform.system() == 'Windows':
        settings_data = settings.open_settings()                                # opening the settings_db.json DB
        full_path_to_MoviesNewRecord = settings_data['path_movie_new_record']
        full_path_to_Movies = settings_data['path_movies_db_sheet']
        os.system(f'start "excel" {full_path_to_Movies}')
        os.system(f'start "excel" {full_path_to_MoviesNewRecord}')