'''
- reorg the 3 diff. routes
- reorg into function what is possible
- decider as a function? re import the modules?
- passing values as function?

'''


from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium import webdriver

from datetime import date

import pyperclip as pc

import sys, webbrowser, platform, shutil


terminal_columns = shutil.get_terminal_size().columns

link = pc.paste()



if platform.system() == 'Windows':
        from openpyxl import load_workbook
        wb = load_workbook('D:/Movies_New_Record.xlsx')
        ws = wb.active

        PATH = 'C:\Program Files (x86)\chromedriver.exe'
        driver = webdriver.Chrome(PATH)
        driver.minimize_window()
        driver.get(link)

if platform.system() == 'Linux':
        from openpyxl import load_workbook
        wb = load_workbook(r'/home/zsandark/Desktop/Movies_New_Record.xlsx')
        ws = wb.active

        PATH = '/home/zsandark/_DEV/Support/Chrome_driver/chromedriver'
        driver = webdriver.Chrome(PATH)
        driver.minimize_window()
        driver.get(link)


try:
        element = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((
                By.CSS_SELECTOR, '.sc-8c396aa2-0 > li:nth-child(1)'))
        )                       
          
        decider = driver.find_element(
        By.CSS_SELECTOR, '.sc-8c396aa2-0 > li:nth-child(1)').text      
except:
        print()
        print('*** ERROR - DECIDER ***')
        print()
        driver.quit()
        sys.exit()

print(decider)

# 1
# len(decider) == 4 -> memoryvie
# 2
# TV Show - TV mini sieres
# 3
# everythin else

driver.quit()
sys.exit()