#!/bin python3.11
'''
TV Show - TV Mini Series
- "Episode Guide" at the top of the page
- under the title, the first section starts with the TYPE of the piece: TV Series 2011–2019 18 57m
'''

from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium import webdriver
from selenium.webdriver.chrome.service import Service

from datetime import date

import pyperclip as pc

import sys, webbrowser, platform, shutil

import f_messages

terminal_columns = shutil.get_terminal_size().columns

# BANNER
f_messages.banner()

link = pc.paste()
cellnumber = 3


# DUAL BOOT OPTION
if platform.system() == 'Windows':
        from openpyxl import load_workbook
        wb = load_workbook('D:/Movies_New_Record.xlsx')
        ws = wb.active

        service = Service('C:\Program Files (x86)\chromedriver.exe')
        driver = webdriver.Chrome(service=service)
        driver.minimize_window()
        driver.get(link)

if platform.system() == 'Linux':
        from openpyxl import load_workbook
        wb = load_workbook(r'/home/zsandark/Desktop/Movies_New_Record.xlsx')
        ws = wb.active

        service = Service('/home/zsandark/_DEV/Support/Chrome_driver/chromedriver')
        driver = webdriver.Chrome(service=service)
        driver.minimize_window()
        driver.get(link)

# TAKING THE VALUES FROM THE 
# MOVIE TITLE - SAME AS MOVIE
try:
        element = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((
                By.CSS_SELECTOR, '.sc-b73cd867-0'))
        )                       
          
        titleRead = driver.find_element(
        By.CSS_SELECTOR, '.sc-b73cd867-0').text
except:
        print()
        print('*** ERROR - MOVIE TITLE ***')
        print()
        driver.quit()
        sys.exit()

# YEAR OF RELEASE - SAME AS MOVIE
try:
        yearRead = driver.find_element(
        By.CSS_SELECTOR, '.sc-8c396aa2-0 > li:nth-child(2) > a:nth-child(1)').text # if only one item(year) there is no index in the last li[1] just li
                   
except:
        print()
        print('*** ERROR - YEAR OF RELEASE ***')
        print()

# DIRECTOR(S) - DIFF!
# directors = []  # instead of director_1_Read - director_3_Read
# try:    
#         for counter in range(1,4):
#                 directors = directors +[driver.find_element(
#                 By.CSS_SELECTOR, f'.sc-fa02f843-0 > ul:nth-child(1) > li:nth-child(1) > div:nth-child(2) > ul:nth-child(1) > li:nth-child({counter}) > a:nth-child(1)').text]
# except:
#         pass # most of the time the movies have only 1 director -> would trigger an error message / not help to identify, if there is a valid error

# STAR(S) - DIFF!
stars = []  # instead star_1_Read - star_3_Read
try:    
        for counter in range(1,4):
                stars = stars + [driver.find_element(
                By.CSS_SELECTOR, f'.sc-fa02f843-0 > ul:nth-child(1) > li:nth-child(2) > div:nth-child(2) > ul:nth-child(1) > li:nth-child({counter}) > a:nth-child(1)').text]
except:
        print()
        print('*** ERROR - STARS ***') # would be triggered if the movie has less than 3 stars
        print()
        
# GENRE(S) - SAME AS MOVIE
genres= [] #instead of genre_1_Read - genre_3_Read
try:    
        for counter in range(1,4):
                genres = genres + [driver.find_element(
        By.CSS_SELECTOR, f'a.sc-16ede01-3:nth-child({counter}) > span:nth-child(1)').text]

except:
        pass # would be triggered if the movie has less than 3 genres

# TAKING THE LENGTH VALUE - DIFF. THAN MOVIE - SAME AS TV MOVIE
try:
        # taking the 2nd item(length) from "2022 1h 33m"
        movieLengthSum = driver.find_element(
        By.CSS_SELECTOR, '.sc-8c396aa2-0 > li:nth-child(3) > a:nth-child(1)').text

        # if the movie has classification(pg-13): "2022 pg-13 1h 33m" taking the 3rd item
        if 'h' not in list(movieLengthSum) or 'm' not in list(movieLengthSum):
                movieLengthSum = driver.find_element(
                By.CSS_SELECTOR, '.sc-8c396aa2-0 > li:nth-child(4)').text
except:
        movieLengthSum = None
        print()
        print('*** ERROR - LENGTH ***')
        print()
       

# VALUATE AND TRANSFORM THE LENGTH VALUE(S)
lengthHour = None
lengthMinute = None
# JUST ONE ITEM LENGTH VALUE, LIKE 45m OR 2h
if len(str(movieLengthSum).split()) == 1:
        if 'h' in list(str(movieLengthSum)):
                lengthHour = str(movieLengthSum).strip('hm')    # removing the "h" or "m" values, i know, in this scenario, just 'h' should be fine
        if 'm' in list(str(movieLengthSum)):
                lengthMinute = str(movieLengthSum).strip('hm')

# TWO ITEMS LENGTH VALUES, LIKE 2h 32m
if len(str(movieLengthSum).split()) == 2:
        lengthList = str(movieLengthSum).split()
        lengthHour = lengthList[0].strip('hm')
        lengthMinute = lengthList[1].strip('hm')


# ADDING THE VALUES TO EXCEL
# MOVIE TITLE

cell = 'C' + str(cellnumber)
ws[cell].value = titleRead
# YEAR OF RELEASE 
cellRYear = 'E' + str(cellnumber)
ws[cellRYear].value = yearRead

# # DIRECTOR(S)
# for counter in range(0,3):
#         cell_director = 'F' + str(int(cellnumber) + counter)
#         try:
#                 ws[cell_director].value = directors[counter]    # adding directors to the sheet (overwriting the privious ones)
#         except:
#                 ws[cell_director].value = None                  # removing previous values, example: the new title has 1 director, the previous one had 3
#                                                                 # the first will be overwritten, the 2nd, 3rd will be removed

# STAR(S)
for counter in range(0,3):
        cell_star = 'G' + str(int(cellnumber) + counter)
        try:
                ws[cell_star].value = stars[counter]
        except:
                ws[cell_star].value = None

# GENRE(S)
genre_columns = ['H', 'I', 'J']
for counter in range(0,3):
        cell_genre = genre_columns[counter] + str(cellnumber)   # writing the genre values horizontally (not vertically like: directors, actors)
        try:
                ws[cell_genre].value = genres[counter]
        except:
                ws[cell_genre].value = None

# MOVIE LENGTH
cellLengthHour = 'Q' + str(cellnumber)
ws[cellLengthHour].value = None                 # removing the previous value from the cell
if lengthHour != None:
        ws[cellLengthHour].value = str(lengthHour)

cellLengthMin = 'R' + str(cellnumber)
ws[cellLengthMin].value = None    
if lengthMinute != None:
        ws[cellLengthMin].value = str(lengthMinute)

# TODAY`S DATE
today = date.today()

day = 'K' + str(cellnumber)
ws[day].value = str(today)[8:]

day = 'L' + str(cellnumber)
ws[day].value = str(today)[5:7]

day = 'M' + str(cellnumber)
ws[day].value = str(today)[0:4]

# HOW MANY TIMES SEEN FORMULA
formula = '=COUNTA(M' + str(cellnumber) + ':M' + str(int(cellnumber) + 2) + ')'  # like: =COUNTA(M6965:M6967)
day = 'N' + str(cellnumber)
ws[day].value = formula

# 1st TIME WATCHING
cellRFirst = 'O' + str(cellnumber)
ws[cellRFirst].value = '1st'        

# SAVE THE SHEET
openSheet = True
while openSheet == True:
        try:
                if platform.system() == 'Windows':
                        wb.save('D:/Movies_New_Record.xlsx')
                        openSheet = False
                        print('\n')

                if platform.system() == 'Linux':
                        wb.save(r'/home/zsandark/Desktop/Movies_New_Record.xlsx')
                        openSheet = False
                        print('\n')
        except:
                print()
                input('!!! ERROR - Close your sheet and hit Enter !!!')

# POSTER IMAGE
try:
        poster = driver.find_element(By.CSS_SELECTOR, '.ipc-media--poster-l > img:nth-child(1)')
        posterLink = poster.get_attribute('src')
        webbrowser.open(posterLink)
except:
        print()
        print('*** ERROR - POSTER ***')
        print()

# LOOKING FOR THE HUNGARIAN TITLE
link = 'https://www.mafab.hu/search/&search='+ ' '.join([titleRead, yearRead])
webbrowser.open(link)

# EPISODE / SERIES LENGTH WARNING
f_messages.tv_show_length()

# BYE BYE BANNER
f_messages.outro()
